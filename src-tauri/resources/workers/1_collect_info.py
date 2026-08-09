#!/usr/bin/env python3
"""FOFA 企业资产候选发现与证据归并。

工作流：
1. discover：使用企业名称、域名、备案、IP、ASN 组织和独有关键词发现候选；
2. expand：只使用人工确认资产的证书/指纹/C 段继续扩展；
3. 输出带证据、评分和人工结论下拉框的复核表。

本程序只调用 FOFA API，不会主动访问或扫描候选资产。
"""

from __future__ import annotations

import argparse
import base64
import configparser
import hashlib
import ipaddress
import json
import logging
import os
import re
import sys
import time
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode, urlparse
from urllib.request import Request, urlopen

import pandas as pd


BASE_DIR = Path(__file__).resolve().parent
FOFA_API_URL = "https://fofa.info/api/v1/search/all"

# FOFA 官方文档标注为基础可返回的字段。扩展字段按会员档位追加。
BASE_FIELDS = [
    "host", "link", "ip", "port", "protocol", "domain", "title", "icp",
    "cert.subject.org", "cert.subject.cn", "cert.domain", "cert.sn", "org",
    "server", "status_code", "jarm",
]
# 当 FOFA 把某个导出字段识别为独立受限功能时，至少保住资产主键和
# Web 基础信息；其余信息可在后续探测阶段补齐。
SAFE_FALLBACK_FIELDS = [
    "host", "link", "ip", "port", "protocol", "domain", "title",
    "icp", "org", "server", "status_code",
]
PROFILE_FIELDS = {
    "basic": [],
    "personal": ["header_hash", "banner_hash", "banner_fid"],
    "professional": [
        "header_hash", "banner_hash", "banner_fid", "cname",
        "lastupdatetime", "product", "product_category",
    ],
    "business": [
        "header_hash", "banner_hash", "banner_fid", "cname",
        "lastupdatetime", "product", "product_category", "product.version",
        "icon_hash", "cname_domain", "cert.is_valid", "cert.is_match",
        "cert.is_equal",
    ],
    "corporate": [
        "header_hash", "banner_hash", "banner_fid", "cname",
        "lastupdatetime", "product", "product_category", "product.version",
        "icon_hash", "cname_domain", "cert.is_valid", "cert.is_match",
        "cert.is_equal", "fid",
    ],
}

OUTPUT_COLUMNS = [
    "company", "decision", "reviewer", "review_note", "score", "confidence",
    "evidence", "phases", "asset_key", *BASE_FIELDS,
    "header_hash", "banner_hash", "banner_fid", "cname", "lastupdatetime",
    "product", "product_category", "product.version", "icon_hash",
    "cname_domain", "cert.is_valid", "cert.is_match", "cert.is_equal", "fid",
    "matched_queries",
]

AFFIRMATIVE_DECISIONS = {"confirmed", "confirm", "yes", "y", "true", "1", "确认", "已确认", "是"}
SPLIT_RE = re.compile(r"[|；;\n]+")
ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def configure_logging(output_dir: Path, verbose: bool = False) -> logging.Logger:
    output_dir.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("fofa_asset_collection")
    logger.setLevel(logging.DEBUG if verbose else logging.INFO)
    logger.handlers.clear()
    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    file_handler = logging.FileHandler(output_dir / "fofa_collection.log", encoding="utf-8")
    stream_handler = logging.StreamHandler()
    file_handler.setFormatter(formatter)
    stream_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    return logger


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    if isinstance(value, (list, tuple, set)):
        if not value:
            return ""
        values = list(value)
        if all(not isinstance(item, (dict, list, tuple, set)) for item in values):
            return " | ".join(clean_value(item) for item in values if clean_value(item))
        return json.dumps(values, ensure_ascii=False, sort_keys=True, default=str)
    if not pd.api.types.is_scalar(value):
        try:
            converted = value.tolist()
        except AttributeError:
            converted = str(value)
        return clean_value(converted)
    if pd.isna(value):
        return ""
    return ILLEGAL_XML_RE.sub("", str(value)).strip()


def nested_value(record: dict[str, Any], dotted_field: str) -> Any:
    """兼容 FOFA 的扁平字段和 cert.subject.org 这类结构化字段。"""
    if dotted_field in record:
        return record[dotted_field]
    value: Any = record
    for part in dotted_field.split("."):
        if not isinstance(value, dict) or part not in value:
            return ""
        value = value[part]
    return value


def flatten_fofa_record(record: dict[str, Any], fields: list[str]) -> dict[str, Any]:
    flattened = {field: nested_value(record, field) for field in fields}

    # FOFA 的结构化结果会把 product/category/version 合并到 product 对象数组。
    products = record.get("product")
    if isinstance(products, list):
        product_names: list[str] = []
        categories: list[str] = []
        versions: list[str] = []
        for item in products:
            if not isinstance(item, dict):
                product_names.append(clean_value(item))
                continue
            if clean_value(item.get("product")):
                product_names.append(clean_value(item.get("product")))
            if clean_value(item.get("category")):
                categories.append(clean_value(item.get("category")))
            if clean_value(item.get("version")):
                versions.append(clean_value(item.get("version")))
        flattened["product"] = " | ".join(dict.fromkeys(filter(None, product_names)))
        flattened["product_category"] = " | ".join(dict.fromkeys(filter(None, categories)))
        flattened["product.version"] = " | ".join(dict.fromkeys(filter(None, versions)))

    return flattened


def split_values(value: Any) -> list[str]:
    return sorted({part.strip() for part in SPLIT_RE.split(clean_value(value)) if part.strip()})


def first_value(row: dict[str, Any], names: Iterable[str]) -> str:
    for name in names:
        if name in row and clean_value(row[name]):
            return clean_value(row[name])
    return ""


def joined_values(row: dict[str, Any], names: Iterable[str]) -> list[str]:
    values: list[str] = []
    for name in names:
        if name in row:
            values.extend(split_values(row[name]))
    return sorted(set(values))


def quote_fofa(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def normalize_host(value: str) -> str:
    value = clean_value(value)
    if not value:
        return ""
    value = value.removeprefix("*.")
    parsed = urlparse(value if "://" in value else f"//{value}")
    return (parsed.hostname or "").removeprefix("*.").rstrip(".").lower()


def root_domain(host: str) -> str:
    """优先用 tldextract 取注册域；不可用时保留原主机，避免错误截断。"""
    host = normalize_host(host)
    if not host:
        return ""
    try:
        ipaddress.ip_address(host)
        return host
    except ValueError:
        pass

    try:
        import tldextract

        extracted = tldextract.TLDExtract(suffix_list_urls=())(host)
        return extracted.top_domain_under_public_suffix or host
    except Exception:
        return host


def public_ipv4_cidr24(value: str) -> str:
    """把明确 IPv4 或 /24 以上小网段归一到 /24；大网段不自动展开。"""
    value = clean_value(value)
    if not value:
        return ""
    try:
        network = ipaddress.ip_network(value, strict=False)
    except ValueError:
        try:
            network = ipaddress.ip_network(f"{value}/32", strict=False)
        except ValueError:
            return ""
    if network.version != 4 or network.prefixlen < 24:
        return ""
    if not network.network_address.is_global:
        return ""
    return str(network.supernet(new_prefix=24) if network.prefixlen > 24 else network)


def read_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        return pd.read_excel(path, dtype=str).fillna("")
    return pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")


def read_lines(path: Path) -> list[str]:
    if not path.exists():
        return []
    return sorted({line.strip() for line in path.read_text(encoding="utf-8-sig").splitlines() if line.strip()})


def load_seeds(seed_path: Path, legacy_domains: Path, legacy_names: Path, logger: logging.Logger) -> list[dict[str, Any]]:
    seeds: list[dict[str, Any]] = []

    if seed_path.exists():
        frame = read_table(seed_path)
        for raw in frame.to_dict("records"):
            names = joined_values(raw, ["names", "full_names", "企业全称", "名称"])
            domains = joined_values(raw, ["domains", "domain", "域名", "urls", "资产域名"])
            icps = joined_values(raw, ["icps", "icp", "备案号"])
            ip_ranges = joined_values(raw, ["ip_ranges", "ips", "ip", "IP段", "IP"])
            asn_orgs = joined_values(raw, ["asn_orgs", "asn_org", "ASN组织", "网络组织"])
            keywords = joined_values(raw, ["keywords", "unique_keywords", "独有关键词", "系统关键词"])
            company = first_value(raw, ["company", "企业", "企业名称", "name"])
            if not company:
                company = next(iter(names or domains or icps or ip_ranges or asn_orgs or keywords), "未命名种子")
            # company is attribution metadata. Only the explicit names column is
            # allowed to create title/body/certificate-organisation queries.
            # This prevents an organisational project label such as “本地” from
            # silently becoming an Internet search condition.
            seeds.append({
                "company": company,
                "names": sorted(set(names)),
                "aliases": joined_values(raw, ["aliases", "简称", "别名"]),
                "domains": domains,
                "icps": icps,
                "ip_ranges": ip_ranges,
                "asn_orgs": asn_orgs,
                "keywords": keywords,
            })
        logger.info("已从 %s 读取 %d 家企业种子", seed_path, len(seeds))
        return seeds

    # 兼容旧版输入。旧文件无法表达“域名属于哪家企业”，所以域名放入单独的未映射分组。
    if not legacy_names.exists():
        for fallback_name in ["assests.txt", "assets.txt"]:
            fallback = BASE_DIR / fallback_name
            if fallback.exists():
                legacy_names = fallback
                logger.info("自动使用现有企业名称文件：%s", fallback)
                break
    names = read_lines(legacy_names)
    domains = read_lines(legacy_domains)
    for name in names:
        seeds.append({
            "company": name, "names": [name], "aliases": [], "domains": [],
            "icps": [], "ip_ranges": [], "asn_orgs": [], "keywords": [],
        })
    if domains:
        seeds.append({
            "company": "未映射域名种子", "names": [], "aliases": [], "domains": domains,
            "icps": [], "ip_ranges": [], "asn_orgs": [], "keywords": [],
        })
    logger.warning("未找到 %s，使用旧版输入；建议尽快改用 seeds.csv 建立企业与种子的映射", seed_path)
    return seeds


@dataclass(frozen=True)
class EvidenceQuery:
    company: str
    phase: str
    category: str
    label: str
    score: int
    query: str


def make_query(company: str, phase: str, category: str, label: str, score: int, expression: str) -> EvidenceQuery:
    return EvidenceQuery(company, phase, category, label, score, expression)


def within_limit(count: int, limit: int) -> bool:
    """0 或负数表示不限制。"""
    return limit <= 0 or count < limit


def build_seed_queries(seeds: list[dict[str, Any]], enable_cidr24: bool, max_cidrs: int, logger: logging.Logger) -> list[EvidenceQuery]:
    queries: list[EvidenceQuery] = []
    cidr_count = 0

    for seed in seeds:
        company = seed["company"]

        for name in seed["names"]:
            q = quote_fofa(name)
            queries.extend([
                make_query(company, "seed", "name_page", f"标题全称:{name}", 45, f"title={q}"),
                make_query(company, "seed", "name_page", f"正文全称:{name}", 30, f"body={q}"),
                make_query(company, "seed", "cert_org", f"证书组织:{name}", 70, f"cert.subject.org={q}"),
            ])

        # 简称通常噪声很大，故默认只作为低分候选。
        for alias in seed["aliases"]:
            q = quote_fofa(alias)
            queries.extend([
                make_query(company, "seed", "alias_page", f"标题简称:{alias}", 20, f"title={q}"),
                make_query(company, "seed", "alias_page", f"正文简称:{alias}", 12, f"body={q}"),
            ])

        for raw_domain in seed["domains"]:
            host = normalize_host(raw_domain)
            if not host:
                continue
            root = root_domain(host)
            queries.append(make_query(company, "seed", "seed_domain", f"精确主机:{host}", 100, f"host={quote_fofa(host)}"))
            queries.append(make_query(company, "seed", "seed_domain", f"注册域:{root}", 95, f"domain={quote_fofa(root)}"))
            queries.append(make_query(company, "seed", "seed_domain", f"证书域:{root}", 90, f"cert.domain={quote_fofa(root)}"))

        for icp in seed["icps"]:
            queries.append(make_query(company, "seed", "icp", f"备案号:{icp}", 95, f"icp={quote_fofa(icp)}"))

        for ip_range in seed["ip_ranges"]:
            queries.append(make_query(company, "seed", "seed_ip", f"明确IP:{ip_range}", 100, f"ip={quote_fofa(ip_range)}"))
            if enable_cidr24 and within_limit(cidr_count, max_cidrs):
                cidr = public_ipv4_cidr24(ip_range)
                if cidr and cidr != ip_range:
                    queries.append(make_query(company, "cidr24", "cidr24", f"种子C段:{cidr}", 12, f"ip={quote_fofa(cidr)}"))
                    cidr_count += 1

        for asn_org in seed["asn_orgs"]:
            queries.append(make_query(company, "seed", "asn_org", f"ASN组织:{asn_org}", 25, f"org={quote_fofa(asn_org)}"))

        for keyword in seed["keywords"]:
            q = quote_fofa(keyword)
            queries.extend([
                make_query(company, "seed", "unique_keyword", f"独有标题:{keyword}", 35, f"title={q}"),
                make_query(company, "seed", "unique_keyword", f"独有正文:{keyword}", 25, f"body={q}"),
            ])

    return deduplicate_queries(queries, logger)


def is_confirmed(row: dict[str, Any], decision_column_exists: bool) -> bool:
    if not decision_column_exists:
        return True
    decision = first_value(row, ["decision", "人工结论", "结论", "confirmed"])
    return decision.strip().lower() in AFFIRMATIVE_DECISIONS


def build_expansion_queries(
    confirmed_path: Path,
    enable_cidr24: bool,
    include_weak: bool,
    max_cidrs: int,
    max_fingerprints: int,
    logger: logging.Logger,
) -> list[EvidenceQuery]:
    if not confirmed_path.exists():
        logger.warning("扩展阶段未找到人工确认文件：%s", confirmed_path)
        return []

    frame = read_table(confirmed_path)
    return build_expansion_queries_from_frame(
        frame, enable_cidr24, include_weak, max_cidrs, max_fingerprints,
        logger, require_confirmed=True,
    )


def build_derived_domain_queries(
    frame: pd.DataFrame,
    seeds: list[dict[str, Any]],
    max_per_company: int,
    logger: logging.Logger,
) -> list[EvidenceQuery]:
    """只从本轮种子证据提取一次域名回查，禁止对回查结果递归扩展。"""
    seed_domains = {normalize_host(domain) for seed in seeds for domain in seed.get("domains", [])}
    domain_only_companies = {
        clean_value(seed.get("company"))
        for seed in seeds
        if seed.get("domains") and not any(seed.get(field) for field in (
            "names", "aliases", "icps", "ip_ranges", "asn_orgs", "keywords",
        ))
    }
    counts: Counter[str] = Counter()
    queries: list[EvidenceQuery] = []
    for row in frame.to_dict("records"):
        company = clean_value(row.get("company"))
        if not company or counts[company] >= max_per_company:
            continue
        # A domain query already returns that registrable domain and its
        # subdomains.  Expanding a domain-only seed through CDN/CNAME evidence
        # turns shared infrastructure (for example dnsv1.com) into a new target
        # and then scans unrelated tenants.
        if company in domain_only_companies:
            continue
        evidence = clean_value(row.get("evidence"))
        cert_org = clean_value(row.get("cert.subject.org"))
        # 域名回查必须有证书组织/明确域名证据，避免公司名称误报扩散。
        if cert_org != company and not any(marker in evidence for marker in ("证书组织:", "精确主机:", "注册域:", "证书域:")):
            continue
        host = normalize_host(clean_value(row.get("domain")) or clean_value(row.get("host")) or clean_value(row.get("link")))
        if not host or host in seed_domains or host.replace("www.", "", 1) in seed_domains:
            continue
        try:
            ipaddress.ip_address(host)
            continue
        except ValueError:
            pass
        root = root_domain(host)
        queries.extend([
            make_query(company, "derived_domain", "derived_host", f"回查主机:{host}", 68, f"host={quote_fofa(host)}"),
            make_query(company, "derived_domain", "derived_domain", f"回查注册域:{root}", 64, f"domain={quote_fofa(root)}"),
            make_query(company, "derived_domain", "derived_cert_domain", f"回查证书域:{root}", 58, f"cert.domain={quote_fofa(root)}"),
        ])
        counts[company] += 1
    logger.info("一次性域名回查生成 %d 条查询（每家公司最多 %d 个派生域名；不递归）", len(queries), max_per_company)
    return deduplicate_queries(queries, logger)


def build_expansion_queries_from_frame(
    frame: pd.DataFrame,
    enable_cidr24: bool,
    include_weak: bool,
    max_cidrs: int,
    max_fingerprints: int,
    logger: logging.Logger,
    require_confirmed: bool,
    auto_min_score: int = 85,
) -> list[EvidenceQuery]:
    decision_column_exists = any(col in frame.columns for col in ["decision", "人工结论", "结论", "confirmed"])
    if require_confirmed and not decision_column_exists:
        logger.warning("确认文件没有 decision/人工结论 列；因文件由 --confirmed 显式传入，将所有行视为已确认")

    queries: list[EvidenceQuery] = []
    cidrs: set[tuple[str, str]] = set()
    fingerprint_count = 0

    for row in frame.to_dict("records"):
        if require_confirmed:
            if not is_confirmed(row, decision_column_exists):
                continue
        else:
            decision = first_value(row, ["decision", "人工结论", "结论", "confirmed"]).strip().lower()
            if decision:
                # 明确 confirmed 可扩展；rejected/uncertain 等其他结论禁止自动扩展。
                if decision not in AFFIRMATIVE_DECISIONS:
                    continue
            else:
                try:
                    score = int(float(clean_value(row.get("score")) or 0))
                except ValueError:
                    score = 0
                if score < auto_min_score:
                    continue
        company = first_value(row, ["company", "企业", "企业名称"])
        if not company:
            continue

        expansion_specs = [
            # 证书序列号可能被共享/CDN复用，单独命中不再直接判定 high。
            ("cert.sn", "cert_serial", "证书序列号", 60),
            ("fid", "fid", "FOFA站点特征", 45),
            ("icon_hash", "icon_hash", "图标哈希", 35),
        ]
        if include_weak:
            # header/banner/JARM 在全网重复度很高，只在显式开启弱指纹时使用。
            expansion_specs.extend([
                ("header_hash", "header_hash", "响应头哈希", 5),
                ("banner_fid", "banner_fid", "Banner结构指纹", 10),
            ])
        for field, category, label, score in expansion_specs:
            value = clean_value(row.get(field, ""))
            if not value or value.lower() in {"nan", "0", "0x0000", "1", "-1"}:
                continue
            if not within_limit(fingerprint_count, max_fingerprints):
                break
            queries.append(make_query(
                company, "fingerprint", category, f"{label}:{value}", score,
                f"{field}={quote_fofa(value)}",
            ))
            fingerprint_count += 1

        if include_weak:
            jarm = clean_value(row.get("jarm", ""))
            if jarm and len(jarm) >= 20 and within_limit(fingerprint_count, max_fingerprints):
                queries.append(make_query(company, "fingerprint", "jarm", f"弱JARM:{jarm}", 8, f"jarm={quote_fofa(jarm)}"))
                fingerprint_count += 1

        if enable_cidr24:
            cidr = public_ipv4_cidr24(clean_value(row.get("ip", "")))
            if cidr:
                cidrs.add((company, cidr))

    selected_cidrs = sorted(cidrs) if max_cidrs <= 0 else sorted(cidrs)[:max_cidrs]
    cidr_source = "确认资产" if require_confirmed else "高可信候选"
    for company, cidr in selected_cidrs:
        queries.append(make_query(company, "cidr24", "cidr24", f"{cidr_source}C段:{cidr}", 12, f"ip={quote_fofa(cidr)}"))

    return deduplicate_queries(queries, logger)


def deduplicate_queries(queries: list[EvidenceQuery], logger: logging.Logger) -> list[EvidenceQuery]:
    unique: dict[tuple[str, str, str], EvidenceQuery] = {}
    for query in queries:
        key = (query.company, query.category, query.query)
        old = unique.get(key)
        if old is None or query.score > old.score:
            unique[key] = query
    result = list(unique.values())
    logger.info("查询计划：原始 %d 条，去重后 %d 条", len(queries), len(result))
    return result


class FofaAPIError(RuntimeError):
    """FOFA 返回的业务错误；与临时网络错误分开，避免无意义重试。"""


class FofaClient:
    def __init__(
        self,
        key: str,
        fields: list[str],
        output_dir: Path,
        interval: float,
        timeout: float,
        use_cache: bool,
        profile_auto_fallback: bool,
        logger: logging.Logger,
    ) -> None:
        self.key = key
        self.fields = fields
        self.output_dir = output_dir
        self.cache_dir = output_dir / "cache"
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self.interval = max(interval, 0.0)
        self.timeout = timeout
        self.use_cache = use_cache
        self.profile_auto_fallback = profile_auto_fallback
        self.auto_profile_resolved = not profile_auto_fallback
        self.force_no_full = False
        self.logger = logger
        self.last_request_at = 0.0
        self.halted = False

    def _cache_path(self, query: str, page: int, size: int, full: bool, fields: list[str]) -> Path:
        material = json.dumps(
            {"query": query, "page": page, "size": size, "full": full, "fields": fields},
            ensure_ascii=False, sort_keys=True,
        )
        return self.cache_dir / f"{hashlib.sha256(material.encode()).hexdigest()}.json"

    def _wait_rate_limit(self) -> None:
        elapsed = time.monotonic() - self.last_request_at
        if elapsed < self.interval:
            time.sleep(self.interval - elapsed)

    def _request(self, query: str, page: int, size: int, full: bool, fields: list[str]) -> dict[str, Any]:
        cache_path = self._cache_path(query, page, size, full, fields)
        if self.use_cache and cache_path.exists():
            return json.loads(cache_path.read_text(encoding="utf-8"))

        qbase64 = base64.b64encode(query.encode("utf-8")).decode("ascii")
        params = {
            "key": self.key,
            "qbase64": qbase64,
            "fields": ",".join(fields),
            "page": page,
            "size": size,
            "full": str(full).lower(),
            "r_type": "json",
        }
        request = Request(
            f"{FOFA_API_URL}?{urlencode(params)}",
            headers={"User-Agent": "Authorized-Asset-Inventory/2.0"},
        )

        last_error: Exception | None = None
        for attempt in range(1, 4):
            try:
                self._wait_rate_limit()
                with urlopen(request, timeout=self.timeout) as response:
                    self.last_request_at = time.monotonic()
                    data = json.loads(response.read().decode("utf-8"))
                if data.get("error"):
                    raise FofaAPIError(clean_value(data.get("errmsg")) or "FOFA 返回未知错误")
                if self.use_cache:
                    cache_path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
                return data
            except HTTPError as exc:
                last_error = RuntimeError(f"HTTP {exc.code}")
                self.last_request_at = time.monotonic()
            except FofaAPIError:
                self.last_request_at = time.monotonic()
                raise
            except (URLError, TimeoutError, ValueError, RuntimeError) as exc:
                last_error = exc
                self.last_request_at = time.monotonic()
            if attempt < 3:
                time.sleep(2 ** (attempt - 1))
        raise RuntimeError(str(last_error))

    def search(self, query: str, page: int, size: int, full: bool) -> tuple[list[dict[str, Any]], int]:
        fields = self.fields
        if self.auto_profile_resolved:
            try:
                data = self._request(query, page, size, False if self.force_no_full else full, fields)
            except FofaAPIError as exc:
                error = str(exc)
                permission_or_limit = any(marker in error for marker in (
                    "820001", "820019", "820041", "824001", "权限", "permission",
                    "已达该功能每日使用上限",
                ))
                if not permission_or_limit:
                    raise
                # 820041 是某个字段/功能的日限额，不等同于整个查询 API
                # 额度耗尽。用旧版兼容的基础字段和普通时间范围重试一次，
                # 成功后本轮固定使用安全字段，避免高级字段阻断所有公司。
                safe_fields = list(dict.fromkeys(SAFE_FALLBACK_FIELDS))
                self.logger.warning("FOFA 当前档位/字段不可用，自动降级为基础字段并关闭全历史后重试：%s", error)
                data = self._request(query, page, size, False, safe_fields)
                self.fields = safe_fields
                self.force_no_full = True
                fields = safe_fields
                self.logger.info("FOFA 安全字段降级成功：本轮后续使用 %d 个基础字段", len(fields))
        else:
            candidates = [
                BASE_FIELDS + PROFILE_FIELDS["business"],
                BASE_FIELDS + PROFILE_FIELDS["professional"],
                BASE_FIELDS + PROFILE_FIELDS["personal"],
                BASE_FIELDS,
            ]
            last_error: Exception | None = None
            data = {}
            for candidate_fields in candidates:
                candidate_fields = list(dict.fromkeys(candidate_fields))
                try:
                    data = self._request(query, page, size, full, candidate_fields)
                    self.fields = candidate_fields
                    fields = candidate_fields
                    self.auto_profile_resolved = True
                    self.logger.info("FOFA 返回字段自动识别完成：%d 个字段", len(fields))
                    break
                except FofaAPIError as exc:
                    last_error = exc
                    self.logger.warning("当前字段档位不可用，尝试较低档位：%s", exc)
            else:
                raise RuntimeError(str(last_error or "无法自动识别 FOFA 字段权限"))

        results: list[dict[str, Any]] = []
        for values in data.get("results", []):
            if isinstance(values, dict):
                results.append(flatten_fofa_record(values, fields))
            else:
                results.append(dict(zip(fields, values)))
        return results, int(data.get("size") or len(results))


class CandidateStore:
    def __init__(self) -> None:
        self.items: dict[tuple[str, str], dict[str, Any]] = {}

    @staticmethod
    def asset_key(record: dict[str, Any]) -> str:
        host = clean_value(record.get("host")) or clean_value(record.get("link"))
        ip = clean_value(record.get("ip"))
        port = clean_value(record.get("port"))
        protocol = clean_value(record.get("protocol"))
        if host:
            return f"{host.lower()}|{ip}|{port}"
        return f"{protocol}|{ip}|{port}"

    def add(self, evidence_query: EvidenceQuery, record: dict[str, Any]) -> None:
        if evidence_query.category == "seed_domain":
            expected = normalize_host(evidence_query.label.rsplit(":", 1)[-1])
            candidate = normalize_host(
                clean_value(record.get("host")) or clean_value(record.get("link"))
            )
            # cert.domain may match one SAN on a shared certificate while the
            # returned service belongs to another tenant.  Explicit domain
            # seeds are strict scope, so the service hostname itself must stay
            # under the supplied registrable domain.
            if expected and candidate and root_domain(candidate) != root_domain(expected):
                return
        asset_key = self.asset_key(record)
        key = (evidence_query.company, asset_key)
        item = self.items.setdefault(key, {
            "company": evidence_query.company,
            "decision": "",
            "reviewer": "",
            "review_note": "",
            "asset_key": asset_key,
            "_scores": {},
            "_evidence": set(),
            "_phases": set(),
            "_queries": set(),
        })
        for field, value in record.items():
            value = clean_value(value)
            if value and not clean_value(item.get(field, "")):
                item[field] = value
        item["_scores"][evidence_query.category] = max(
            evidence_query.score,
            item["_scores"].get(evidence_query.category, 0),
        )
        item["_evidence"].add(evidence_query.label)
        item["_phases"].add(evidence_query.phase)
        item["_queries"].add(evidence_query.query)

    def to_frame(self) -> pd.DataFrame:
        rows = []
        for item in self.items.values():
            score = min(100, sum(item["_scores"].values()))
            if score >= 85:
                confidence = "high"
            elif score >= 55:
                confidence = "medium"
            else:
                confidence = "low"
            row = {column: clean_value(item.get(column, "")) for column in OUTPUT_COLUMNS}
            row.update({
                "score": score,
                "confidence": confidence,
                "evidence": " | ".join(sorted(item["_evidence"])),
                "phases": " | ".join(sorted(item["_phases"])),
                "matched_queries": " | ".join(sorted(item["_queries"])),
            })
            rows.append(row)
        if not rows:
            return pd.DataFrame(columns=OUTPUT_COLUMNS)
        return pd.DataFrame(rows, columns=OUTPUT_COLUMNS).sort_values(
            ["company", "score", "ip", "port"], ascending=[True, False, True, True]
        )


def merge_existing_review(frame: pd.DataFrame, review_path: Path, logger: logging.Logger) -> pd.DataFrame:
    """保留上次人工结论，并合并本轮新增证据，防止扩展运行覆盖复核成果。"""
    if not review_path.exists():
        return frame
    try:
        previous = read_table(review_path)
    except Exception as exc:
        logger.warning("读取既有人工复核表失败，不执行结论继承：%s", exc)
        return frame
    if previous.empty:
        return frame

    current_records: dict[tuple[str, str], dict[str, Any]] = {}
    for raw in frame.to_dict("records"):
        row = {column: clean_value(raw.get(column, "")) for column in OUTPUT_COLUMNS}
        key = (row["company"], row["asset_key"] or CandidateStore.asset_key(row))
        current_records[key] = row

    for raw in previous.to_dict("records"):
        old = {column: clean_value(raw.get(column, "")) for column in OUTPUT_COLUMNS}
        old["asset_key"] = old["asset_key"] or CandidateStore.asset_key(old)
        key = (old["company"], old["asset_key"])
        current = current_records.get(key)
        if current is None:
            current_records[key] = old
            continue

        # 人工字段始终以既有表为准。
        for field in ["decision", "reviewer", "review_note"]:
            if old[field]:
                current[field] = old[field]

        for field in ["evidence", "phases", "matched_queries"]:
            merged = sorted({part.strip() for value in [old[field], current[field]] for part in value.split(" | ") if part.strip()})
            current[field] = " | ".join(merged)

        try:
            current_score = int(float(current["score"] or 0))
        except ValueError:
            current_score = 0
        try:
            old_score = int(float(old["score"] or 0))
        except ValueError:
            old_score = 0
        current["score"] = max(current_score, old_score)
        current["confidence"] = "high" if current["score"] >= 85 else "medium" if current["score"] >= 55 else "low"

    normalized_rows = []
    for row in current_records.values():
        try:
            score = int(float(clean_value(row.get("score")) or 0))
        except ValueError:
            score = 0
        row["score"] = score
        row["confidence"] = "high" if score >= 85 else "medium" if score >= 55 else "low"
        normalized_rows.append(row)
    merged_frame = pd.DataFrame(normalized_rows, columns=OUTPUT_COLUMNS)
    merged_frame = merged_frame.sort_values(
        ["company", "score", "ip", "port"], ascending=[True, False, True, True]
    )
    logger.info("已继承既有复核表：合并后 %d 条，人工结论不会被覆盖", len(merged_frame))
    return merged_frame


def execute_queries(
    queries: list[EvidenceQuery],
    client: FofaClient,
    store: CandidateStore,
    page_size: int,
    max_pages: int,
    full: bool,
    logger: logging.Logger,
) -> pd.DataFrame:
    query_log: list[dict[str, Any]] = []
    total_queries = len(queries)
    consecutive_service_errors = 0
    consecutive_feature_limits = 0

    for index, evidence_query in enumerate(queries, 1):
        if client.halted:
            logger.error("采集器已触发保护停机，剩余 %d 条查询留待下次从缓存续跑", total_queries - index + 1)
            break
        # FOFA 官方限制：查询表达式包含 body 时，size 最大 500。
        effective_size = min(page_size, 500) if "body=" in evidence_query.query else page_size
        fetched = 0
        status = "ok"
        error = ""
        total_matches = 0
        logger.info(
            "[%d/%d] %s | %s | %s",
            index, total_queries, evidence_query.company, evidence_query.label, evidence_query.query,
        )
        page = 1
        while True:
            try:
                records, total_matches = client.search(evidence_query.query, page, effective_size, full)
            except Exception as exc:
                status = "error"
                error = str(exc)
                logger.error("查询失败：%s", exc)
                if any(marker in error.lower() for marker in (
                    "connection refused", "timed out", "temporary failure",
                    "name or service not known", "nodename nor servname",
                    "network is unreachable",
                )):
                    client.halted = True
                    status = "fatal_network"
                    logger.error("FOFA 网络/代理不可达，立即停止剩余查询；请先在配置中心运行 FOFA 连通性测试")
                if "[820031]" in error or "F点余额不足" in error:
                    client.halted = True
                    status = "fatal_fpoints"
                    logger.error("检测到 FOFA 全局额度/F点不足，立即停止后续查询；本轮不会伪装成成功")
                elif "820041" in error or "已达该功能每日使用上限" in error:
                    consecutive_feature_limits += 1
                    status = "skipped_feature_limit"
                    logger.warning(
                        "当前查询触发 FOFA 单项功能日限额，跳过该条件并继续基础查询（连续 %d 次）",
                        consecutive_feature_limits,
                    )
                    if consecutive_feature_limits >= 3:
                        client.halted = True
                        status = "fatal_feature_guard"
                        logger.error("连续 3 个条件均触发单项功能日限额，保护停机，避免无效消耗")
                elif "[-501]" in error or "服务错误" in error or "HTTP 429" in error:
                    consecutive_service_errors += 1
                    if consecutive_service_errors >= 3:
                        client.halted = True
                        status = "fatal_service_guard"
                        logger.error("连续 %d 次服务/限流错误，触发保护停机", consecutive_service_errors)
                break
            consecutive_service_errors = 0
            consecutive_feature_limits = 0
            for record in records:
                store.add(evidence_query, record)
            fetched += len(records)
            if len(records) < effective_size or fetched >= total_matches:
                break
            if max_pages > 0 and page >= max_pages:
                break
            page += 1

        query_log.append({
            "company": evidence_query.company,
            "phase": evidence_query.phase,
            "category": evidence_query.category,
            "label": evidence_query.label,
            "score": evidence_query.score,
            "query": evidence_query.query,
            "total_matches": total_matches,
            "fetched": fetched,
            "status": status,
            "error": error,
        })

    return pd.DataFrame(query_log)


def select_manual_review_candidates(frame: pd.DataFrame) -> pd.DataFrame:
    """把完整候选分成可人工处理的三档；弱指纹全集仍保留在 candidates.csv。"""
    if frame.empty:
        result = frame.copy()
        result.insert(min(6, len(result.columns)), "review_tier", [])
        return result

    evidence = frame["evidence"].fillna("").astype(str)
    company = frame["company"].fillna("").astype(str).str.strip()
    cert_org = frame["cert.subject.org"].fillna("").astype(str).str.strip()
    protocol = frame["protocol"].fillna("").astype(str).str.lower()
    updated = frame["lastupdatetime"].fillna("").astype(str)
    status = frame["status_code"].fillna("").astype(str)

    cert_exact = company.ne("") & cert_org.eq(company)
    name_match = evidence.str.contains("标题全称:|正文全称:|独有标题:|独有正文:", regex=True)
    is_web = protocol.str.contains("http", regex=False)
    recent = updated.ge("2025-01-01")
    active = status.isin({"200", "201", "202", "204", "206", "301", "302", "303", "307", "308", "401", "403"})

    p1 = cert_exact & is_web & recent & active
    p2 = cert_exact & is_web & recent & ~active
    p3 = ~cert_exact & name_match & is_web & recent & active
    selected = p1 | p2 | p3
    result = frame.loc[selected].copy()
    result.insert(6, "review_tier", "")
    result.loc[p1[selected], "review_tier"] = "P1_强归属且近期活跃"
    result.loc[p2[selected], "review_tier"] = "P2_强归属但入口待验证"
    result.loc[p3[selected], "review_tier"] = "P3_名称命中待确认"
    return result.sort_values(["review_tier", "company", "score"], ascending=[True, True, False])


def write_review_workbook(frame: pd.DataFrame, path: Path, logger: logging.Logger) -> None:
    try:
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            frame.to_excel(writer, index=False, sheet_name="manual_review")
            worksheet = writer.book["manual_review"]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            decision_col = frame.columns.get_loc("decision") + 1
            score_col = frame.columns.get_loc("score") + 1
            decision_letter = worksheet.cell(1, decision_col).column_letter
            score_letter = worksheet.cell(1, score_col).column_letter

            validation = DataValidation(
                type="list", formula1='"confirmed,rejected,uncertain"', allow_blank=True
            )
            worksheet.add_data_validation(validation)
            validation.add(f"{decision_letter}2:{decision_letter}{max(2, len(frame) + 1)}")

            green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            worksheet.conditional_formatting.add(
                f"{score_letter}2:{score_letter}{max(2, len(frame) + 1)}",
                CellIsRule(operator="greaterThanOrEqual", formula=["85"], fill=green),
            )
            worksheet.conditional_formatting.add(
                f"{score_letter}2:{score_letter}{max(2, len(frame) + 1)}",
                CellIsRule(operator="between", formula=["55", "84"], fill=yellow),
            )
            worksheet.conditional_formatting.add(
                f"{score_letter}2:{score_letter}{max(2, len(frame) + 1)}",
                CellIsRule(operator="lessThan", formula=["55"], fill=red),
            )

            widths = {
                "company": 24, "decision": 14, "reviewer": 12, "review_note": 30,
                "score": 9, "confidence": 12, "evidence": 60, "host": 40,
                "link": 40, "ip": 18, "port": 9, "title": 35,
                "matched_queries": 60,
            }
            for column_index, column_name in enumerate(frame.columns, 1):
                worksheet.column_dimensions[worksheet.cell(1, column_index).column_letter].width = widths.get(column_name, 18)
        logger.info("人工复核表已写入：%s", path)
    except Exception as exc:
        logger.warning("Excel 复核表写入失败，仅保留 CSV：%s", exc)


def write_seed_template(path: Path) -> None:
    if path.exists():
        return
    frame = pd.DataFrame([{
        "company": "示例企业（请删除本行）",
        "names": "示例企业有限公司|示例企业集团有限公司",
        "aliases": "示例企业",
        "domains": "example.com|example.cn",
        "icps": "京ICP备00000000号",
        "ip_ranges": "203.0.113.10|203.0.113.0/24",
        "asn_orgs": "只有明确知道ASN组织名时填写",
        "keywords": "独有系统名称|独有SSO标识",
    }])
    frame.to_csv(path, index=False, encoding="utf-8-sig")


def load_config(path: Path) -> configparser.ConfigParser:
    config = configparser.ConfigParser(interpolation=None)
    if path.exists():
        config.read(path, encoding="utf-8")
    return config


def parse_args() -> argparse.Namespace:
    pre_parser = argparse.ArgumentParser(add_help=False)
    pre_parser.add_argument("--config", type=Path, default=BASE_DIR / "config.ini")
    pre_args, _ = pre_parser.parse_known_args()
    config = load_config(pre_args.config)
    collection = config["collection"] if config.has_section("collection") else {}

    def cfg(name: str, default: Any) -> Any:
        return collection.get(name, default) if collection else default

    def cfg_int(name: str, default: int) -> int:
        try:
            return int(cfg(name, default))
        except (TypeError, ValueError):
            return default

    def cfg_float(name: str, default: float) -> float:
        try:
            return float(cfg(name, default))
        except (TypeError, ValueError):
            return default

    def cfg_bool(name: str, default: bool) -> bool:
        value = str(cfg(name, str(default))).strip().lower()
        return value in {"1", "true", "yes", "on", "y", "是"}

    parser = argparse.ArgumentParser(description="FOFA 企业资产候选发现、指纹扩展和人工确认表生成")
    parser.add_argument("--config", type=Path, default=pre_args.config, help="账号、Key 和默认采集参数 INI")
    parser.add_argument("--seeds", type=Path, default=BASE_DIR / "seeds.csv", help="企业种子 CSV/XLSX")
    parser.add_argument("--legacy-domains", type=Path, default=BASE_DIR / "original_urls.txt")
    parser.add_argument("--legacy-names", type=Path, default=BASE_DIR / "names.txt")
    configured_output = Path(str(cfg("output_dir", "output")))
    if not configured_output.is_absolute():
        configured_output = BASE_DIR / configured_output
    configured_confirmed = Path(str(cfg("confirmed", configured_output / "manual_review.xlsx")))
    if not configured_confirmed.is_absolute():
        configured_confirmed = BASE_DIR / configured_confirmed
    parser.add_argument("--confirmed", type=Path, default=configured_confirmed, help="带 confirmed 结论的复核表")
    parser.add_argument("--output-dir", type=Path, default=configured_output)
    parser.add_argument("--mode", choices=["discover", "expand", "all"], default=cfg("mode", "all"))
    parser.add_argument("--profile", choices=["auto", *PROFILE_FIELDS], default=cfg("profile", "auto"), help="FOFA会员字段档位")
    parser.add_argument("--page-size", type=int, default=cfg_int("page_size", 500))
    parser.add_argument("--max-pages", type=int, default=cfg_int("max_pages", 0), help="每条查询最大页数；0=全部分页")
    parser.add_argument("--interval", type=float, default=cfg_float("interval", 2.0))
    parser.add_argument("--timeout", type=float, default=cfg_float("timeout", 30.0))
    parser.add_argument("--full", action=argparse.BooleanOptionalAction, default=cfg_bool("full", True), help="查询 FOFA 全部历史数据")
    parser.add_argument("--enable-cidr24", action=argparse.BooleanOptionalAction, default=cfg_bool("enable_cidr24", True), help="对种子IP/高可信候选/人工确认IP进行 /24 扩展")
    parser.add_argument("--max-cidrs", type=int, default=cfg_int("max_cidrs", 0), help="C段查询上限；0=不限制")
    parser.add_argument("--max-fingerprints", type=int, default=cfg_int("max_fingerprints", 0), help="指纹查询上限；0=不限制")
    parser.add_argument("--max-derived-domains", type=int, default=cfg_int("max_derived_domains", 200), help="每家公司从本轮证据派生回查的域名上限；0=不限制")
    parser.add_argument("--include-weak-fingerprints", action=argparse.BooleanOptionalAction, default=cfg_bool("include_weak_fingerprints", True), help="启用低唯一性的 JARM 扩展")
    parser.add_argument("--auto-expand-min-score", type=int, default=cfg_int("auto_expand_min_score", 85), help="all 模式自动扩展候选的最低分")
    parser.add_argument("--cache", action=argparse.BooleanOptionalAction, default=cfg_bool("cache", True), help="使用本地响应缓存")
    parser.add_argument("--verbose", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    logger = configure_logging(args.output_dir, args.verbose)
    config = load_config(args.config)
    key = os.environ.get("FOFA_KEY", "").strip() or config.get("fofa", "key", fallback="").strip()
    if not key:
        logger.error("未找到 FOFA Key；请填写 %s 的 [fofa] key，或设置 FOFA_KEY", args.config)
        return 2
    if args.config.exists():
        logger.info("已读取本地配置：%s（Key 不会写入日志）", args.config)

    if args.profile == "auto":
        fields = BASE_FIELDS + PROFILE_FIELDS["business"]
        auto_fallback = True
    else:
        fields = BASE_FIELDS + PROFILE_FIELDS[args.profile]
        auto_fallback = False
    fields = list(dict.fromkeys(fields))

    args.output_dir.mkdir(parents=True, exist_ok=True)
    all_queries: list[EvidenceQuery] = []
    query_logs: list[pd.DataFrame] = []
    store = CandidateStore()
    client = FofaClient(
        key=key,
        fields=fields,
        output_dir=args.output_dir,
        interval=args.interval,
        timeout=args.timeout,
        use_cache=args.cache,
        profile_auto_fallback=auto_fallback,
        logger=logger,
    )
    logger.info(
        "FOFA 请求配置：endpoint=%s | profile=%s | full=%s | page_size=%d | max_pages=%d | fields=%s",
        FOFA_API_URL,
        args.profile,
        args.full,
        args.page_size,
        args.max_pages,
        ",".join(fields),
    )

    seed_queries: list[EvidenceQuery] = []
    discovered_seeds: list[dict[str, Any]] = []
    if args.mode in {"discover", "all"}:
        seeds = load_seeds(args.seeds, args.legacy_domains, args.legacy_names, logger)
        if not seeds:
            template = args.seeds.with_name("seeds.example.csv")
            write_seed_template(template)
            logger.error("没有可用种子。已生成模板：%s", template)
            return 2
        discovered_seeds = seeds
        seed_queries = build_seed_queries(seeds, args.enable_cidr24, args.max_cidrs, logger)
        all_queries.extend(seed_queries)
        query_logs.append(execute_queries(
            seed_queries, client, store, max(1, args.page_size), args.max_pages,
            args.full, logger,
        ))

    # 只对种子查询得到的强证据域名回查一次；回查结果不会再次产生域名回查。
    expansion_base = store.to_frame()
    if discovered_seeds and not expansion_base.empty and args.mode in {"discover", "all"}:
        derived_queries = build_derived_domain_queries(
            expansion_base, discovered_seeds, args.max_derived_domains, logger,
        )
        if derived_queries:
            all_queries.extend(derived_queries)
            query_logs.append(execute_queries(
                derived_queries, client, store, max(1, args.page_size), args.max_pages,
                args.full, logger,
            ))

    expansion_queries: list[EvidenceQuery] = []
    if args.mode == "expand":
        expansion_queries = build_expansion_queries(
            args.confirmed, args.enable_cidr24, args.include_weak_fingerprints,
            args.max_cidrs, args.max_fingerprints, logger,
        )
    elif args.mode == "all":
        # 同一次运行先完成种子发现，再从既有人工确认资产和本轮高可信候选提取指纹/C段。
        # expansion 使用域名回查前的快照，避免“回查→再提取→再回查”的隐性循环。
        provisional = merge_existing_review(expansion_base, args.confirmed, logger)
        domain_only_companies = {
            clean_value(seed.get("company"))
            for seed in discovered_seeds
            if seed.get("domains") and not any(seed.get(field) for field in (
                "names", "aliases", "icps", "ip_ranges", "asn_orgs", "keywords",
            ))
        }
        if domain_only_companies and not provisional.empty:
            before = len(provisional)
            provisional = provisional[
                ~provisional["company"].map(clean_value).isin(domain_only_companies)
            ]
            logger.info(
                "域名-only 种子启用严格范围：跳过 %d 条共享指纹/C段扩展候选",
                before - len(provisional),
            )
        expansion_queries = build_expansion_queries_from_frame(
            provisional,
            args.enable_cidr24,
            args.include_weak_fingerprints,
            args.max_cidrs,
            args.max_fingerprints,
            logger,
            require_confirmed=False,
            auto_min_score=args.auto_expand_min_score,
        )
        executed = {(query.company, query.category, query.query) for query in seed_queries}
        expansion_queries = [
            query for query in expansion_queries
            if (query.company, query.category, query.query) not in executed
        ]

    if expansion_queries:
        all_queries.extend(expansion_queries)
        query_logs.append(execute_queries(
            expansion_queries, client, store, max(1, args.page_size), args.max_pages,
            args.full, logger,
        ))

    if client.halted:
        logger.warning("本轮因额度/服务保护提前停止；已完成结果和缓存仍会正常保存")

    all_queries = deduplicate_queries(all_queries, logger)
    if not all_queries:
        logger.error("没有生成任何查询任务")
        return 2

    plan_frame = pd.DataFrame([query.__dict__ for query in all_queries])
    plan_frame.to_csv(args.output_dir / "query_plan.csv", index=False, encoding="utf-8-sig")
    query_log = pd.concat(query_logs, ignore_index=True) if query_logs else pd.DataFrame()
    query_log.to_csv(args.output_dir / "query_log.csv", index=False, encoding="utf-8-sig")

    candidates = store.to_frame()
    if client.halted and candidates.empty:
        logger.error("FOFA 未返回候选资产，本轮任务应重试前先检查额度、权限和配置；未产生可入库数据")
        return 3
    if candidates.empty:
        logger.warning("本轮查询完成但候选资产为 0：可能是确无匹配，也可能是查询权限/字段与当前 FOFA 账户不符")
    candidates = merge_existing_review(candidates, args.output_dir / "manual_review.xlsx", logger)
    candidates.to_csv(args.output_dir / "candidates.csv", index=False, encoding="utf-8-sig")
    review_candidates = select_manual_review_candidates(candidates)
    review_candidates.to_csv(args.output_dir / "review_candidates.csv", index=False, encoding="utf-8-sig")
    write_review_workbook(review_candidates, args.output_dir / "manual_review.xlsx", logger)

    logger.info(
        "完成：候选 %d 条（high=%d, medium=%d, low=%d）",
        len(candidates),
        int((candidates["confidence"] == "high").sum()) if not candidates.empty else 0,
        int((candidates["confidence"] == "medium").sum()) if not candidates.empty else 0,
        int((candidates["confidence"] == "low").sum()) if not candidates.empty else 0,
    )
    logger.info("下一步：在 manual_review.xlsx 的 decision 列标记 confirmed/rejected/uncertain")
    logger.info("完整候选保留 %d 条；进入人工复核 %d 条", len(candidates), len(review_candidates))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
